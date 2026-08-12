#!/usr/bin/env python3
"""
Generate the downloadable import-template workbooks (plan tasks 6+7).

Reads the canonical vCard field snapshot (fixtures/vcard-fields.snapshot.json)
and writes two .xlsx files:

  import-template-horizontal.xlsx — row 1 = the 30 canonical field headers
    (snake_case ids), row 2 = one synthetic example contact; more contacts go
    in rows 3, 4, ...
  import-template-vertical.xlsx — column A = the 30 headers, column B = the
    example values; more contacts go in columns C, D, ...

Both are committed as static assets at front-cards/public/templates/ and are
served to both Demo and Normal mode. Run inside the api-server container
(openpyxl is installed there):

    python3 /app/batch-parsing/generate_import_templates.py --outdir /tmp/templates
    python3 /app/batch-parsing/generate_import_templates.py --verify /tmp/templates

--verify re-opens both workbooks and asserts the headers match the snapshot
exactly (order + values) and the example row/column matches EXAMPLE_VALUES.
"""

import argparse
import json
import os
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font

HORIZONTAL_FILENAME = "import-template-horizontal.xlsx"
VERTICAL_FILENAME = "import-template-vertical.xlsx"

DEFAULT_SNAPSHOT = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "fixtures", "vcard-fields.snapshot.json"
)

INSTRUCTION = (
    "Headers use the canonical snake_case field names; English/Spanish alias "
    "names (e.g. 'Full name', 'Nombre completo') are also accepted."
)

# Synthetic example values keyed by canonical field id — obviously fake, no PII.
EXAMPLE_VALUES = {
    "full_name": "Example Person",
    "first_name": "Example",
    "last_name": "Person",
    "work_phone": "+506 2222 0000",
    "work_phone_ext": "123",
    "mobile_phone": "+506 8888 0000",
    "email": "example.person@example.com",
    "address_street": "123 Example Street",
    "address_city": "Example City",
    "address_state": "Example State",
    "address_postal": "10101",
    "address_country": "Example Country",
    "social_instagram": "example.person",
    "social_twitter": "@exampleperson",
    "social_facebook": "example.person",
    "business_name": "Example Company S.A.",
    "business_title": "Example Job Title",
    "business_department": "Example Department",
    "business_url": "https://example.com",
    "business_hours": "Mon-Fri 8:00-17:00",
    "business_address_street": "456 Business Avenue",
    "business_address_city": "Business City",
    "business_address_state": "Business State",
    "business_address_postal": "20202",
    "business_address_country": "Business Country",
    "business_linkedin": "https://linkedin.com/company/example",
    "business_twitter": "@examplecompany",
    "personal_url": "https://example-person.example.com",
    "personal_bio": "Example bio text.",
    "personal_birthday": "1990-01-01",
}


def load_field_ids(snapshot_path: str) -> list:
    """Canonical field ids in snapshot order."""
    with open(snapshot_path, encoding="utf-8") as f:
        fields = json.load(f)
    ids = [entry["id"] for entry in fields]
    missing = [fid for fid in ids if fid not in EXAMPLE_VALUES]
    if missing:
        raise ValueError(f"EXAMPLE_VALUES missing entries for snapshot fields: {missing}")
    return ids


def generate(outdir: str, snapshot_path: str = DEFAULT_SNAPSHOT) -> list:
    """Write both workbooks into outdir; returns the generated paths."""
    ids = load_field_ids(snapshot_path)
    os.makedirs(outdir, exist_ok=True)

    horizontal_path = os.path.join(outdir, HORIZONTAL_FILENAME)
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    for col, fid in enumerate(ids, start=1):
        cell = ws.cell(row=1, column=col, value=fid)
        cell.font = Font(bold=True)
        ws.cell(row=2, column=col, value=EXAMPLE_VALUES[fid])
    ws.cell(row=1, column=1).comment = Comment(
        f"{INSTRUCTION} One contact per row, starting at row 2.", "import-template"
    )
    wb.save(horizontal_path)

    vertical_path = os.path.join(outdir, VERTICAL_FILENAME)
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    for row, fid in enumerate(ids, start=1):
        cell = ws.cell(row=row, column=1, value=fid)
        cell.font = Font(bold=True)
        ws.cell(row=row, column=2, value=EXAMPLE_VALUES[fid])
    ws.cell(row=1, column=1).comment = Comment(
        f"{INSTRUCTION} One contact per column, starting at column B.", "import-template"
    )
    wb.save(vertical_path)

    return [horizontal_path, vertical_path]


def verify(outdir: str, snapshot_path: str = DEFAULT_SNAPSHOT) -> None:
    """Assert both workbooks in outdir match the snapshot headers (order +
    values) and the EXAMPLE_VALUES example contact. Raises AssertionError."""
    ids = load_field_ids(snapshot_path)
    expected_examples = [EXAMPLE_VALUES[fid] for fid in ids]

    wb = load_workbook(os.path.join(outdir, HORIZONTAL_FILENAME))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(ids) + 1)]
    assert ws.max_column == len(ids), (
        f"{HORIZONTAL_FILENAME}: expected {len(ids)} columns, found {ws.max_column}"
    )
    assert headers == ids, f"{HORIZONTAL_FILENAME}: header row mismatch: {headers}"
    examples = [ws.cell(row=2, column=c).value for c in range(1, len(ids) + 1)]
    assert examples == expected_examples, (
        f"{HORIZONTAL_FILENAME}: example row mismatch: {examples}"
    )

    wb = load_workbook(os.path.join(outdir, VERTICAL_FILENAME))
    ws = wb.active
    assert ws.max_row == len(ids), (
        f"{VERTICAL_FILENAME}: expected {len(ids)} rows, found {ws.max_row}"
    )
    headers = [ws.cell(row=r, column=1).value for r in range(1, len(ids) + 1)]
    assert headers == ids, f"{VERTICAL_FILENAME}: header column mismatch: {headers}"
    examples = [ws.cell(row=r, column=2).value for r in range(1, len(ids) + 1)]
    assert examples == expected_examples, (
        f"{VERTICAL_FILENAME}: example column mismatch: {examples}"
    )


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--snapshot",
        default=DEFAULT_SNAPSHOT,
        help="Path to vcard-fields.snapshot.json (default: fixtures/ next to this script)",
    )
    parser.add_argument(
        "--outdir",
        default=".",
        help="Directory to write the workbooks into (default: current directory)",
    )
    parser.add_argument(
        "--verify",
        metavar="DIR",
        help="Verify the workbooks in DIR against the snapshot instead of generating",
    )
    args = parser.parse_args(argv)

    if args.verify:
        verify(args.verify, args.snapshot)
        print(f"OK: {HORIZONTAL_FILENAME} and {VERTICAL_FILENAME} match the snapshot")
        return 0

    for path in generate(args.outdir, args.snapshot):
        print(f"wrote {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
